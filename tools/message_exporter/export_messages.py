#!/usr/bin/env python3
"""Convert message XLSX data into the JSON format used by Godot."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - friendly CLI error for missing dependency
    load_workbook = None


ALLOWED_POLICIES = {"queue", "interrupt", "ignore"}
DEFAULTS: dict[str, Any] = {
    "duration": 3.0,
    "typewriter": True,
    "chars_per_second": 24,
    "priority": 0,
    "interrupt_policy": "queue",
}
REQUIRED_COLUMNS = [
    "id",
    "speaker",
    "text",
    "duration",
    "typewriter",
    "chars_per_second",
    "priority",
    "interrupt_policy",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export message XLSX to Godot JSON.")
    parser.add_argument("--input", required=True, help="Input XLSX file path.")
    parser.add_argument("--output", required=True, help="Output JSON file path.")
    parser.add_argument("--sheet", default=None, help="Optional worksheet name. Defaults to active sheet.")
    return parser.parse_args()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_bool(value: Any, default: bool) -> bool:
    if isinstance(value, bool):
        return value
    normalized = cell_text(value).lower()
    if normalized == "":
        return default
    if normalized in {"true", "1", "yes", "y"}:
        return True
    if normalized in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"Invalid boolean value: {value}")


def parse_float(value: Any, default: float) -> float:
    if cell_text(value) == "":
        return default
    return float(value)


def parse_int(value: Any, default: int) -> int:
    if cell_text(value) == "":
        return default
    return int(value)


def get_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, value in enumerate(header_row):
        column_name = cell_text(value)
        if column_name != "":
            header_map[column_name] = index

    missing = [name for name in REQUIRED_COLUMNS if name not in header_map]
    if missing:
        raise ValueError(f"Missing XLSX columns: {', '.join(missing)}")
    return header_map


def row_value(row: tuple[Any, ...], header_map: dict[str, int], key: str) -> Any:
    index = header_map[key]
    if index >= len(row):
        return None
    return row[index]


def build_message(row: tuple[Any, ...], header_map: dict[str, int], row_number: int) -> tuple[str, dict[str, Any]]:
    message_id = cell_text(row_value(row, header_map, "id"))
    text = cell_text(row_value(row, header_map, "text"))
    if message_id == "":
        raise ValueError(f"Row {row_number}: id cannot be empty")
    if text == "":
        raise ValueError(f"Row {row_number}: text cannot be empty")

    interrupt_policy = cell_text(row_value(row, header_map, "interrupt_policy")).lower() or DEFAULTS["interrupt_policy"]
    if interrupt_policy not in ALLOWED_POLICIES:
        raise ValueError(f"Row {row_number}: invalid interrupt_policy '{interrupt_policy}'")

    message = {
        "speaker": cell_text(row_value(row, header_map, "speaker")),
        "text": text,
        "duration": parse_float(row_value(row, header_map, "duration"), DEFAULTS["duration"]),
        "typewriter": parse_bool(row_value(row, header_map, "typewriter"), DEFAULTS["typewriter"]),
        "chars_per_second": parse_int(row_value(row, header_map, "chars_per_second"), DEFAULTS["chars_per_second"]),
        "priority": parse_int(row_value(row, header_map, "priority"), DEFAULTS["priority"]),
        "interrupt_policy": interrupt_policy,
    }
    return message_id, message


def load_messages(input_path: Path, sheet_name: str | None) -> dict[str, dict[str, Any]]:
    if load_workbook is None:
        raise ValueError("openpyxl is required to read XLSX files. Install it with: pip install openpyxl")
    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("Input file must be .xlsx")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as error:
            raise ValueError("XLSX file is empty") from error

        header_map = get_header_map(header_row)
        messages: dict[str, dict[str, Any]] = {}
        for row_number, row in enumerate(rows, start=2):
            if all(cell_text(value) == "" for value in row):
                continue
            message_id, message = build_message(row, header_map, row_number)
            if message_id in messages:
                raise ValueError(f"Row {row_number}: duplicate id '{message_id}'")
            messages[message_id] = message
        return messages
    finally:
        workbook.close()


def write_json(output_path: Path, messages: dict[str, dict[str, Any]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="\n") as json_file:
        json.dump(messages, json_file, ensure_ascii=False, indent=2)
        json_file.write("\n")


def main() -> int:
    args = parse_args()
    try:
        messages = load_messages(Path(args.input), args.sheet)
        write_json(Path(args.output), messages)
    except (OSError, ValueError, KeyError) as error:
        print(f"error: {error}", file=sys.stderr)
        return 1

    print(f"exported {len(messages)} messages to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
